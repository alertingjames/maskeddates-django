import datetime
import difflib
import os
import string
import urllib
from itertools import islice

import io
import requests
import xlrd
import re

from django.core import mail
from django.core.mail import send_mail, BadHeaderError, EmailMessage
from django.contrib import messages
from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives

from django.core.files.storage import FileSystemStorage
import json
from django.contrib import auth
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.utils.datastructures import MultiValueDictKeyError
from django.views.decorators.cache import cache_control
from numpy import long

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny
from xlrd import XLRDError
from time import gmtime, strftime
import time
from openpyxl.styles import PatternFill

from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User, AnonymousUser
from django.contrib.auth import login, authenticate
from django.conf import settings
from django import forms
import sys
from django.core.cache import cache
# import unirest
from dating.models import Member, Post, Notification, Unravel, Dynamical, Plan, Bid, Folder, File, Picture, Note, Settings
from dating.serializers import MemberSerializer, PostSerializer, NotificationSerializer, UnravelSerializer, \
    DynamicalSerializer, PlanSerializer, PictureSerializer

imei_encoded = {
    0:'a',
    1:'d',
    2:'x',
    3:'p',
    4:'s',
    5:'j',
    6:'o',
    7:'z',
    8:'c',
    9:'w'
}

imei_decoded = {
    'a':'0',
    'd':'1',
    'x':'2',
    'p':'3',
    's':'4',
    'j':'5',
    'o':'6',
    'z':'7',
    'c':'8',
    'w':'9'
}

def index(request):
    return HttpResponse('<h2>Hello Masked Dates!</h2>')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def register_member(request):

    if request.method == 'POST':

        eml = request.POST.get('email', '')
        name = request.POST.get('name', '')
        password = request.POST.get('password', '')
        fbphoto = request.POST.get('fb_photo', '')
        gender = request.POST.get('gender', '')
        age = request.POST.get('age', '')
        phone = request.POST.get('phone', '')
        address = request.POST.get('location', '')
        lat = request.POST.get('lat', '')
        lng = request.POST.get('lng', '')
        answers = request.POST.get('answers', '')
        answers2 = request.POST.get('answers2', '')
        premium = request.POST.get('premium', '')
        photos = request.POST.get('photos', '')
        photo_unlock = request.POST.get('photo_unlock', '')
        selfie_approved = request.POST.get('selfie_approved', '')
        lastlogin = request.POST.get('last_login', '')
        actives = request.POST.get('actives', '')
        phone_imei = request.POST.get('phone_imei', '')

        ss = ''
        if len(phone_imei) > 0:
            for i in range(0, len(phone_imei)):
                ss = ss + imei_encoded[int(phone_imei[i])]

        members = Member.objects.filter(phone_imei=ss)
        if members.count() > 0:
            member = members[0]
            if member.password == '' or member.password == 'activated':
                email = member.email
                resp = {'result_code': '100', 'email': email}
                return JsonResponse(resp)

        if actives == '':
            actives = '0'

        if photo_unlock == '':
            photo_unlock = '0'

        users = Member.objects.filter(email=eml)
        count = users.count()

        if count == 0:
            member = Member()
            member.email = eml
            member.name = name
            member.password = password
            member.fb_photo = fbphoto
            member.gender = gender
            member.age = age
            member.phone = phone
            member.address = address
            member.lat = lat
            member.lng = lng
            member.answers = answers
            member.answers2 = answers2

            member.premium = premium
            member.photos = photos
            member.photo_unlock = photo_unlock
            member.selfie_approved = selfie_approved
            member.last_login = lastlogin
            member.actives = actives
            member.phone_imei = ss

            member.save()

            resp = {'result_code': '0', 'member_id': member.pk}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '101'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload__member_picture(request):

    if request.method == 'POST':

        image = request.FILES['file']

        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)

        member_id = request.POST.get('member_id')
        member = Member.objects.get(id=member_id)
        member.photo_url = settings.URL + uploaded_file_url
        member.selfie_approved = ''
        member.save()

        notis = Notification.objects.filter(sender_id=member_id)
        if notis.count() > 0:
            for noti in notis:
                noti.sender_photo = settings.URL + uploaded_file_url
                noti.save()

        resp = {'result_code': '0', 'photo_url': member.photo_url}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_member(request):

    if request.method == 'POST':
        name = request.POST.get('name', '')
        eml = request.POST.get('email', '')
        gender = request.POST.get('gender', '')
        age = request.POST.get('age', '')
        address = request.POST.get('location', '')
        lat = request.POST.get('lat', '')
        lng = request.POST.get('lng', '')
        phone_imei = request.POST.get('phone_imei', '')

        users = Member.objects.filter(email=eml)
        count = users.count()
        if count > 0:
            member = users[0]
            member.name = name
            member.gender = gender
            member.age = age
            member.address = address
            member.lat = lat
            member.lng = lng
            member.phone_imei = phone_imei

            member.save()

            notis = Notification.objects.filter(sender_email=eml)
            if notis.count() > 0:
                for noti in notis:
                    noti.sender_name = name
                    noti.save()

            resp = {'result_code': '0', 'member_id': member.pk}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)
        else:
            resp_er = {'result_code': '101'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def member_deactivate(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        member = Member.objects.get(id=member_id)
        member.password = 'canceled'
        member.save()
        resp = {
            'result_code':'0'
        }
        return JsonResponse(resp)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def register_location(request):

    if request.method == 'POST':

        eml = request.POST.get('email', '')
        address = request.POST.get('address', '')
        lat = request.POST.get('lat', '')
        lng = request.POST.get('lng', '')

        users = Member.objects.filter(email=eml)
        count = users.count()

        if count > 0:
            member = users[0]
            member.address = address
            member.lat = lat
            member.lng = lng

            member.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '101'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_member(request):

    if request.method == 'POST':

        email = request.POST.get('email', '')
        phone_imei = request.POST.get('phone_imei', '')

        ss = ''
        if len(phone_imei) > 0:
            for i in range(0, len(phone_imei)):
                ss = ss + imei_encoded[int(phone_imei[i])]

        members = Member.objects.filter(phone_imei=ss)
        if members.count() > 0:
            member = members[0]
            if member.password == '' or member.password == 'activated':
                if member.email != email:
                    resp = {'result_code': '100', 'email': member.email}
                    return JsonResponse(resp)

        users = Member.objects.filter(email=email)
        count = users.count()
        if count > 0:
            user = users[0]
            user.last_login = str(int(round(time.time() * 1000)))
            user.save()
            serializer = MemberSerializer(users, many=True)
            resp = {'result_code': '0', 'data': serializer.data}

            return JsonResponse(resp, status=status.HTTP_200_OK)

        else:
            resp = {'result_code': '1'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_member2(request):

    if request.method == 'POST':

        email = request.POST.get('email', '')
        fb_photo = request.POST.get('fb_photo', '')

        phone_imei = request.POST.get('phone_imei', '')

        ss = ''
        if len(phone_imei) > 0:
            for i in range(0, len(phone_imei)):
                ss = ss + imei_encoded[int(phone_imei[i])]

        members = Member.objects.filter(phone_imei=ss)
        if members.count() > 0:
            member = members[0]
            if member.password == '' or member.password == 'activated':
                if member.email != email:
                    resp = {'result_code': '100', 'email': member.email}
                    return JsonResponse(resp)

        users = Member.objects.filter(email=email)
        count = users.count()
        if count > 0:
            user = users[0]
            user.fb_photo = fb_photo
            user.last_login = str(int(round(time.time() * 1000)))
            user.save()
            serializer = MemberSerializer(users, many=True)
            resp = {'result_code': '0', 'data': serializer.data}

            return JsonResponse(resp, status=status.HTTP_200_OK)

        else:
            resp = {'result_code': '1'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload_questionnaire(request):

    if request.method == 'POST':

        eml = request.POST.get('email', '')
        questionnaires = request.POST.get('questionnaireinfo', '')

        users = Member.objects.filter(email=eml)
        count = users.count()

        if count > 0:
            member = users[0]
            member.answers = questionnaires

            member.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '101'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_all_users(request):
    userList = []
    premiumList = []
    contacted_premiumList = []
    alluserList = []
    allpremiumList = []
    contactedList = []
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        users = Member.objects.all().order_by('-id')
        ###################################################################################################################

        list = []
        for p in range(0, users.count()):
            userp = users[p]
            max_last_login = int(userp.last_login)
            for q in range(p, users.count()):
                userq = users[q]
                if int(userq.last_login) >= max_last_login and not userq in list:
                    if userq.password != 'deactivated' and userq.password != 'canceled':
                        list.insert(0, userq)

        #########################################################################################################################
        for user in list:
            notifications = Notification.objects.filter(user_id=user_id, sender_id=user.pk)
            notifications2 = Notification.objects.filter(sender_id=user_id, user_id=user.pk)
            if notifications.count() == 0 and notifications2.count() == 0:
                if len(user.premium) > 0:
                    premiumList.append(user)
                    allpremiumList.append(user)
                    continue
                userList.append(user)
                alluserList.append(user)
            else:
                if len(user.premium) > 0:
                    contacted_premiumList.append(user)
                    allpremiumList.append(user)
                    continue
                contactedList.append(user)
                alluserList.append(user)

        x = 0
        y = 0

        allList = []

        if len(allpremiumList) > 0:
            for i in range(0, len(allpremiumList)):
                x += 1
                if x < 3:
                    allList.append(allpremiumList[i])
                    try:
                        if allpremiumList[i + 1] is None and len(alluserList) > 0:
                            for j in range(y, len(alluserList)):
                                allList.append(alluserList[j])
                    except IndexError:
                        for j in range(y, len(alluserList)):
                            allList.append(alluserList[j])
                else:
                    x = 0
                    try:
                        if alluserList[y] is not None:
                            allList.append(alluserList[y])
                            y += 1
                            try:
                                if allpremiumList[i + 1] is None and len(alluserList) > 0:
                                    for j in range(y, len(alluserList)):
                                        allList.append(alluserList[j])
                            except IndexError:
                                for j in range(y, len(alluserList)):
                                    allList.append(alluserList[j])
                    except IndexError:
                        print ('IndexError')
        else:
            for j in range(0, len(alluserList)):
                allList.append(alluserList[j])

        x = 0
        y = 0

        newList = []

        if len(premiumList) > 0:
            for i in range(0, len(premiumList)):
                x += 1
                if x < 3:
                    newList.append(premiumList[i])
                    try:
                        if premiumList[i + 1] is None and len(userList) > 0:
                            for j in range(y, len(userList)):
                                newList.append(userList[j])
                    except IndexError:
                        for j in range(y, len(userList)):
                            newList.append(userList[j])
                else:
                    x = 0
                    try:
                        if userList[y] is not None:
                            newList.append(userList[y])
                            y += 1
                            try:
                                if premiumList[i + 1] is None and len(userList) > 0:
                                    for j in range(y, len(userList)):
                                        newList.append(userList[j])
                            except IndexError:
                                for j in range(y, len(userList)):
                                    newList.append(userList[j])
                    except IndexError:
                        print ('IndexError')
        else:
            for j in range(0, len(userList)):
                newList.append(userList[j])

        x = 0
        y = 0

        contList = []

        if len(contacted_premiumList) > 0:
            for i in range(0, len(contacted_premiumList)):
                x += 1
                if x < 3:
                    contList.append(contacted_premiumList[i])
                    try:
                        if contacted_premiumList[i + 1] is None and len(contactedList) > 0:
                            for j in range(y, len(contactedList)):
                                contList.append(contactedList[j])
                    except IndexError:
                        for j in range(y, len(contactedList)):
                            contList.append(contactedList[j])
                else:
                    x = 0
                    try:
                        if contactedList[y] is not None:
                            contList.append(contactedList[y])
                            y += 1
                            try:
                                if contacted_premiumList[i + 1] is None and len(contactedList) > 0:
                                    for j in range(y, len(contactedList)):
                                        contList.append(contactedList[j])
                            except IndexError:
                                for j in range(y, len(contactedList)):
                                    contList.append(contactedList[j])
                    except IndexError:
                        print ('IndexError')
        else:
            for j in range(0, len(contactedList)):
                contList.append(contactedList[j])


        serializer_new = MemberSerializer(newList, many=True)
        serializer_all = MemberSerializer(allList, many=True)
        serializer_contacted = MemberSerializer(contList, many=True)

        resp = {'result_code': '0', 'users': serializer_new.data, 'allusers': serializer_all.data, 'contacted': serializer_contacted.data}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload_questionnaire2(request):

    if request.method == 'POST':

        eml = request.POST.get('email', '')
        questionnaires = request.POST.get('questionnaireinfo', '')

        users = Member.objects.filter(email=eml)
        count = users.count()

        if count > 0:
            member = users[0]
            member.answers2 = questionnaires

            member.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '101'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def post(request):

    if request.method == 'POST':

        eml = request.POST.get('email', '')
        postText = request.POST.get('post', '')

        users = Member.objects.filter(email=eml)
        count = users.count()

        if count > 0:
            member = users[0]
            post = Post()
            post.user_id = member.pk
            post.photo = ''
            post.text = postText
            post.datetime = str(int(round(time.time() * 1000)))

            post.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '1'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updatePost(request):

    if request.method == 'POST':

        id = request.POST.get('id', '')
        postText = request.POST.get('post', '')

        post = Post.objects.get(id=id)
        post.text = postText
        post.datetime = str(int(round(time.time() * 1000)))

        post.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_posts(request):

    if request.method == 'POST':

        email = request.POST.get('email', '')
        users = Member.objects.filter(email=email)
        count = users.count()

        if count > 0:

            posts = Post.objects.filter(user_id=users[0].pk).order_by('-id')

            serializer = PostSerializer(posts, many=True)
            resp = {'result_code': '0', 'posts': serializer.data, }
            return JsonResponse(resp, status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '1'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delpost(request):
    if request.method == 'POST':
        idx = request.POST.get('post_id', '')
        Post.objects.get(id=idx).delete()
        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendNotification(request):

    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        sender_id = request.POST.get('sender_id', '')
        name = request.POST.get('sender_name', '')
        email = request.POST.get('sender_email', '')
        photo = request.POST.get('sender_photo', '')
        notitext = request.POST.get('notitext', '')
        notitime = request.POST.get('notitime', '')
        option = request.POST.get('option', '')

        if option == 'accepted' or option == 'declined' or option == 'blocked':
            notis2 = Notification.objects.filter(user_id=sender_id, sender_id=user_id)
            noti2 = notis2[0]
            noti2.option = option
            if option == 'blocked':
                noti2.option = 'block'
                noti2.active = '0'
            elif option == 'accepted':
                noti2.unraveled = '1'
                noti2.enteredtime = str(int(round(time.time() * 1000)))
            elif option == 'declined':
                noti2.unraveled = '0'
            noti2.save()

        notis = Notification.objects.filter(user_id=user_id, sender_id=sender_id)

        if notis.count() == 0:
            noti = Notification()
            noti.user_id = user_id
            noti.sender_id = sender_id
            noti.sender_name = name
            noti.sender_email = email
            noti.sender_photo = photo
            noti.notitext = notitext
            noti.notitime = notitime
            if option == 'requested':
                noti.unraveled = '0'
                noti.active = '0'
            elif option == 'accepted':
                noti.unraveled = '1'
                noti.enteredtime = str(int(round(time.time() * 1000)))
                noti.active = '0'
            elif option == 'declined':
                noti.unraveled = '0'
                noti.active = '0'
            noti.option = option
            noti.exitedtime = ''
            noti.status = ''

            noti.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            noti = notis[0]
            noti.user_id = user_id
            noti.sender_id = sender_id
            noti.sender_name = name
            noti.sender_email = email
            noti.sender_photo = photo
            noti.notitext = notitext
            noti.notitime = notitime
            if option == 'requested':
                noti.unraveled = '0'
            elif option == 'blocked':
                noti.active = '0'
            elif option == 'declined':
                noti.unraveled = '0'
                noti.active = '0'
            noti.option = option

            noti.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getNotifications(request):

    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        notis = Notification.objects.filter(user_id=user_id)
        count = notis.count()

        serializer = NotificationSerializer(notis, many=True)
        resp = {'result_code': '0', 'data': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getUser(request):

    if request.method == 'POST':

        user_id = request.POST.get('user_id', '')

        user = Member.objects.get(id=user_id)
        users = Member.objects.filter(email=user.email)

        serializer = MemberSerializer(users, many=True)
        resp = {'result_code': '0', 'users': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def checkNotification(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        notis = Notification.objects.filter(user_id=user_id, sender_id=me_id)
        notis2 = Notification.objects.filter(user_id=me_id, sender_id=user_id)
        count = notis.count()
        count2 = notis2.count()
        if count > 0 or count2 > 0:
            sts = ''
            try:
                sts = notis[0].option
            except IndexError:
                sts = notis2[0].option
            resp = {'result_code': '0', 'status':sts}
            return JsonResponse(resp)
        else:
            resp = {'result_code': '1'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def removeNotification(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        notis = Notification.objects.filter(user_id=me_id, sender_id=user_id)
        count = notis.count()
        if count > 0:
            notification = notis[0]
            notification.delete()
            resp = {'result_code': '0'}
            return JsonResponse(resp)
        else:
            resp = {'result_code': '1'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def checkUnravel(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        unravels = Unravel.objects.filter(user_id=user_id, me_id=me_id)
        if unravels.count() > 0:
            serializer = UnravelSerializer(unravels, many=True)
            resp = {'result_code': '0', 'unravels': serializer.data}
            return JsonResponse(resp)
        else:
            unravel = Unravel()
            unravel.me_id = me_id
            unravel.user_id = user_id
            unravel.unraveled = '0'
            unravel.active = '0'
            unravel.entered_datetime = str(int(round(time.time() * 1000)))
            unravel.exited_datetime = '0'
            unravel.save()

            unravels = Unravel.objects.filter(user_id=user_id, me_id=me_id)
            serializer = UnravelSerializer(unravels, many=True)
            resp = {'result_code': '0', 'unravels': serializer.data}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getActives(request):
    if request.method == 'POST':
        me_id = request.POST.get('me_id', '')
        unravels = Unravel.objects.filter(me_id=me_id)
        serializer = UnravelSerializer(unravels, many=True)
        resp = {'result_code': '0', 'unravels': serializer.data}
        return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def setActive(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        notis = Notification.objects.filter(user_id=me_id, sender_id=user_id)
        if notis.count() > 0:
            notification = notis[0]
            if notification.option != 'block':
                notification.active = '1'
            notification.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)
    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def riseUnraveled(request):

    if request.method == 'POST':

        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        notis = Notification.objects.filter(user_id=me_id, sender_id=user_id)
        if notis.count() > 0:
            notification = notis[0]
            if int(notification.unraveled) < 4:
                notification.unraveled = str(int(notification.unraveled) + 1)
                notification.enteredtime = str(int(round(time.time() * 1000)))
                notification.save()

        notifications = Notification.objects.filter(user_id=user_id, sender_id=me_id)
        if notifications.count() > 0:
            notification = notifications[0]
            if int(notification.unraveled) < 4:
                notification.unraveled = str(int(notification.unraveled) + 1)
                notification.enteredtime = str(int(round(time.time() * 1000)))
                notification.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def exitChat(request):
    if request.method == 'POST':
        me_id = request.POST.get('me_id', '')

        notis = Notification.objects.filter(user_id=me_id)
        if notis.count() > 0:
            for notification in notis:
                notification.active = '0'
                notification.exitedtime = str(int(round(time.time() * 1000)))
                notification.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def leaveLastMessage(request):
    if request.method == 'POST':
        receiver_id = request.POST.get('receiver_id', '')
        sender_id = request.POST.get('sender_id', '')
        last_message = request.POST.get('last_message', '')

        notis = Notification.objects.filter(user_id=receiver_id, sender_id=sender_id)
        notis2 = Notification.objects.filter(user_id=sender_id, sender_id=receiver_id)
        if notis.count() > 0:
            noti = notis[0]
            noti.notitext = last_message
            noti.save()

        if notis2.count() > 0:
            noti2 = notis2[0]
            noti2.notitext = last_message
            noti2.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def chatExit(request):

    if request.method == 'POST':
        me_id = request.POST.get('me_id', '')
        user_id = request.POST.get('user_id', '')

        notis = Notification.objects.filter(user=me_id, sender_id=user_id)
        if notis.count() > 0:
            notification = notis[0]
            notification.active = '0'
            notification.exitedtime = str(int(round(time.time() * 1000)))
            notification.save()
        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)
    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def setNoti(request):

    if request.method == 'POST':
        notiJsonStr = request.POST.get('notifications', None)

        try:
            decoded = json.loads(notiJsonStr)
            for notiJson in decoded['notijson']:

                noti_id = notiJson['noti_id']
                notitext = notiJson['notitext']
                datetime = notiJson['datetime']

                noti = Notification.objects.get(id=noti_id)
                noti.notitext = notitext
                noti.notitime = datetime
                noti.save()

            resp = {'result_code': '0'}
            return JsonResponse(resp, status=status.HTTP_200_OK)

        except:
            resp = {'result_code': '1'}
            return JsonResponse(resp)
    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getNoti(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        notis = Notification.objects.filter(user_id=me_id, sender_id=user_id)
        serializer = NotificationSerializer(notis, many=True)
        resp = {'result_code': '0', 'noti': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)
    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upgradeUser(request):

    if request.method == 'POST':

        user_id = request.POST.get('user_id', '')
        expired = request.POST.get('expired_time', '')

        user = Member.objects.get(id=user_id)
        user.premium = expired
        user.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def resetPremium(request):

    if request.method == 'POST':

        user_id = request.POST.get('me_id', '')

        user = Member.objects.get(id=user_id)
        user.premium = ''
        user.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getActiveDates(request):
    datesList = []
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')

        notis = Notification.objects.filter(user_id=user_id)
        count = notis.count()
        if count > 0:
            for noti in notis:
                if noti.active == '1' or noti.option == 'block':
                    datesList.insert(0, noti)

        serializer = NotificationSerializer(datesList, many=True)
        resp = {'result_code': '0', 'data': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def releaseBlock(request):

    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        notis2 = Notification.objects.filter(user_id=me_id, sender_id=user_id)
        noti2 = notis2[0]
        noti2.option = 'normal'
        noti2.save()

        notis = Notification.objects.filter(user_id=user_id, sender_id=me_id)

        noti = notis[0]
        noti.option = 'normal'
        noti.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def releaseBlockOnChat(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id', '')
        me_id = request.POST.get('me_id', '')

        notis2 = Notification.objects.filter(user_id=me_id, sender_id=user_id)
        noti2 = notis2[0]
        noti2.option = 'normal'
        noti2.active = '1'
        noti2.save()

        notis = Notification.objects.filter(user_id=user_id, sender_id=me_id)

        noti = notis[0]
        noti.option = 'normal'
        noti.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

def adminloginpage(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return render(request, 'dating/login.html')
    else:
        return redirect('/home')

def logout(request):
    auth.logout(request)
    cache.clear()
    request.session.flush()
    request.user = AnonymousUser
    for key in request.session.keys():
        del request.session[key]
    return redirect('/')

def loginAdmin(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        photo = request.POST.get('photo', '')

        # users = User.objects.filter(email=email)
        # if users.count() == 0:
        #     user1 = User()
        #     user1.username = email
        #     user1.email = email
        #     user1.password = 'admin12345'
        #     user1.set_password('admin12345')
        #     user1.save()

        user = authenticate(username=email, password='admin12345')
        if user is not None:
            login(request, user)
            members = Member.objects.all().order_by('-id')
            members = getmembers(members)
            return render(request, 'dating/home.html', {'result':'success', 'users':members})
        else:
            return render(request, 'dating/result.html', {'result': 'You don\'t have any permission to access this website'})
    else:return redirect('/home')

def view_image(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    image = request.GET['photo_url']
    return render(request, 'dating/view_picture.html', {'image':image})

def photo_approved(request, member_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    member = Member.objects.get(id=member_id)
    member.selfie_approved = 'yes'
    member.save()
    members = Member.objects.all().order_by('-id')
    members = getmembers(members)
    return render(request, 'dating/home.html', {'member_id': member_id, 'users': members})

def home(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    request.session['filter'] = 'all_users'
    members = Member.objects.all().order_by('-id')
    members = getmembers(members)
    return render(request, 'dating/home.html', {'users': members, 'filterF':'true'})

def photo_rejected(request, member_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    member = Member.objects.get(id=member_id)
    member.selfie_approved = 'no'
    member.save()
    members = Member.objects.all().order_by('-id')
    members = getmembers(members)
    return render(request, 'dating/home.html', {'member_id': member_id, 'users': members})

def questions(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    dynams = Dynamical.objects.all().order_by('-id')
    return render(request, 'dating/dynamical.html', {'questions':dynams})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_dynamical(request):
    if request.method == 'POST':
        text = request.POST.get('text', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return render(request, 'dating/login.html')
        qs = Dynamical.objects.filter(dynamical_question=text)
        context = {}
        if qs.count() == 0:
            q = Dynamical()
            q.dynamical_question = text
            q.save()
            qs = Dynamical.objects.all().order_by('-id')
            context = {'questions':qs, 'note':'update'}
        else:
            qs = Dynamical.objects.all().order_by('-id')
            context = {'questions': qs, 'note':'exist'}
        return render(request, 'dating/dynamical.html', context)

    else:
        return redirect('/questions')

def delete_dynamical(request, q_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    q = Dynamical.objects.get(id=q_id)
    if q is not None:
        q.delete()
    qs = Dynamical.objects.all().order_by('-id')
    context = {'questions': qs, 'note': 'update'}
    return render(request, 'dating/dynamical.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getDynamicals(request):
    qs = Dynamical.objects.all()
    serializer = DynamicalSerializer(qs, many=True)
    resp = {'result_code':'0', 'data':serializer.data}
    return JsonResponse(resp, status=status.HTTP_200_OK)

def editmembership(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    plans = Plan.objects.all()
    plan = None
    if plans.count() == 0:
        plan = Plan()
        plan.months = ''
        plan.dates = '3'
        plan.price = ''
        plan.months1 = '3'
        plan.dates1 = '6'
        plan.price1 = '21.99'
        plan.months2 = '6'
        plan.dates2 = '14'
        plan.price2 = '27.99'
        plan.months3 = '12'
        plan.dates3 = '28'
        plan.price3 = '42.99'
        plan.save()
    else:
        plan = plans[0]
    return render(request, 'dating/membership.html', {'plan':plan})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updatemembership(request):
    if request.method == 'POST':
        months = request.POST.get('months', '')
        dates = request.POST.get('dates', '3')
        price = request.POST.get('price', '')

        months1 = request.POST.get('months1', '1')
        dates1 = request.POST.get('dates1', '3')
        price1 = request.POST.get('price1', '0')

        months2 = request.POST.get('months2', '1')
        dates2 = request.POST.get('dates2', '3')
        price2 = request.POST.get('price2', '0')

        months3 = request.POST.get('months3', '1')
        dates3 = request.POST.get('dates3', '3')
        price3 = request.POST.get('price3', '0')

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return render(request, 'dating/login.html')

        plans = Plan.objects.all()
        if plans.count() > 0:
            plan = plans[0]
            plan.months = months
            plan.dates = dates
            plan.price = price
            plan.months1 = months1
            plan.dates1 = dates1
            plan.price1 = price1
            plan.months2 = months2
            plan.dates2 = dates2
            plan.price2 = price2
            plan.months3 = months3
            plan.dates3 = dates3
            plan.price3 = price3
            plan.save()
        return redirect('/editmembership')

    else:
        return redirect('/editmembership')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getmembership(request):
    if request.method == 'POST':
        plans = Plan.objects.all()
        serializer = PlanSerializer(plans, many=True)
        resp = {'result_code':'0', 'data':serializer.data}
        return JsonResponse(resp)

def getmembers(members):
    memberList = []
    for member in members:
        mbr = Member()
        mbr.id = member.pk
        mbr.name = member.name
        mbr.email = member.email
        mbr.age = member.age
        mbr.gender = member.gender
        mbr.address = member.address
        mbr.photo_url = member.photo_url
        mbr.fb_photo = member.fb_photo
        mbr.password = member.password
        mbr.phone = member.phone
        mbr.selfie_approved = member.selfie_approved
        premiumCaption = 'Bronze member'
        plans = Plan.objects.all()
        plan = plans[0]
        premium_max_dates_list = [0, plan.dates1, plan.dates2, plan.dates3]
        if member.premium:
            premiumStr = member.premium
            expired, num = premiumStr.split("_")
            expired = time.strftime('%Y-%m-%d', time.gmtime(int(expired) / 1000.0))
            if num == '1': premiumCaption = 'Silver member'
            elif num == '2': premiumCaption = 'Gold member'
            elif num == '3': premiumCaption = 'Platinum member'
            else: num == '0'
            mbr.premium = premiumCaption + ' till ' + expired + ', Max dates: ' + str(premium_max_dates_list[int(num)])
        else: mbr.premium = premiumCaption + ', Max dates: ' + str(plan.dates)
        memberList.append(mbr)
    return memberList

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def filter(request):
    if request.method == 'POST':
        fromdate = request.POST.get('fromdate', '')
        todate = request.POST.get('todate', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return render(request, 'dating/login.html')
        notis = Notification.objects.all()
        users = Member.objects.all().order_by('-id')
        userList = []
        for noti in notis:
            if noti.enteredtime >= fromdate and noti.enteredtime <= todate:
                for user in users:
                    if int(noti.user_id) == user.pk or int(noti.sender_id) == user.pk:
                        if not user in userList:
                            userList.append(user)
        users = getmembers(userList)
        return render(request, 'dating/home.html', {'users':users})

    else: return redirect('/home')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def search(request):
    if request.method == 'POST':
        q = request.POST.get('q', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return render(request, 'dating/login.html')
        users = Member.objects.filter(name__icontains=q).order_by('-id')
        users = getmembers(users)
        if len(users) > 0:
            return render(request, 'dating/home.html', {'users': users})
        else:
            users = Member.objects.filter(email__icontains=q).order_by('-id')
            users = getmembers(users)
            if len(users) > 0:
                return render(request, 'dating/home.html', {'users': users})
            else:
                users = Member.objects.filter(age__icontains=q).order_by('-id')
                users = getmembers(users)
                if len(users) > 0:
                    return render(request, 'dating/home.html', {'users': users})
                else:
                    users = Member.objects.filter(address__icontains=q).order_by('-id')
                    users = getmembers(users)
                    if len(users) > 0:
                        return render(request, 'dating/home.html', {'users': users})
                    else:
                        users = Member.objects.filter(gender__icontains=q).order_by('-id')
                        users = getmembers(users)
                        if len(users) > 0:
                            return render(request, 'dating/home.html', {'users': users})
                        else:
                            users = Member.objects.filter(answers__icontains=q).order_by('-id')
                            users = getmembers(users)
                            if len(users) > 0:
                                return render(request, 'dating/home.html', {'users': users})
                            else:
                                users = Member.objects.filter(answers2__icontains=q).order_by('-id')
                                users = getmembers(users)
                                return render(request, 'dating/home.html', {'users': users})
    else:
        return redirect('/home')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def ctrsearch(request):
    if request.method == 'POST':
        q = request.POST.get('q', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return render(request, 'dating/login.html')
        plan = Plan.objects.all()[0]
        users = Member.objects.filter(name__icontains=q).order_by('-id')
        users = getmembers(users)
        if len(users) > 0:
            return render(request, 'dating/control.html', {'users': users, 'plan':plan})
        else:
            users = Member.objects.filter(email__icontains=q).order_by('-id')
            users = getmembers(users)
            if len(users) > 0:
                return render(request, 'dating/control.html', {'users': users, 'plan':plan})
            else:
                users = Member.objects.filter(age__icontains=q).order_by('-id')
                users = getmembers(users)
                if len(users) > 0:
                    return render(request, 'dating/control.html', {'users': users, 'plan':plan})
                else:
                    users = Member.objects.filter(address__icontains=q).order_by('-id')
                    users = getmembers(users)
                    if len(users) > 0:
                        return render(request, 'dating/control.html', {'users': users, 'plan':plan})
                    else:
                        users = Member.objects.filter(gender__icontains=q).order_by('-id')
                        users = getmembers(users)
                        if len(users) > 0:
                            return render(request, 'dating/control.html', {'users': users, 'plan':plan})
                        else:
                            users = Member.objects.filter(answers__icontains=q).order_by('-id')
                            users = getmembers(users)
                            if len(users) > 0:
                                return render(request, 'dating/control.html', {'users': users, 'plan':plan})
                            else:
                                users = Member.objects.filter(answers2__icontains=q).order_by('-id')
                                users = getmembers(users)
                                return render(request, 'dating/control.html', {'users': users, 'plan':plan})
    else:
        return redirect('/control')

def new_photo_users(request):
    request.session['filter'] = 'new_users'
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    users = Member.objects.all().order_by('-id')
    userList = []
    for user in users:
        if user.selfie_approved == '':
            userList.append(user)
    users = getmembers(userList)
    return render(request, 'dating/home.html', {'users':users, 'filterF':'true'})

def photo_approved_users(request):
    request.session['filter'] = 'approved_users'
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    users = Member.objects.all().order_by('-id')
    userList = []
    for user in users:
        if user.selfie_approved == 'yes':
            userList.append(user)
    users = getmembers(userList)
    return render(request, 'dating/home.html', {'users': users, 'filterF':'true'})

def photo_rejected_users(request):
    request.session['filter'] = 'rejected_users'
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    users = Member.objects.all().order_by('-id')
    userList = []
    for user in users:
        if user.selfie_approved == 'no':
            userList.append(user)
    users = getmembers(userList)
    return render(request, 'dating/home.html', {'users': users, 'filterF':'true'})

def export_xlsx(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    filtered = request.session['filter']
    users = Member.objects.all().order_by('-id')
    userList = []
    docName = 'Filtered Users'
    if filtered == 'all_users':
        docName = 'All Users'
        for user in users:
            userList.append(user)
    elif filtered == 'new_users':
        docName = 'Photo-New Users'
        for user in users:
            if user.selfie_approved == '':
                userList.append(user)
    elif filtered == 'approved_users':
        docName = 'Photo-Approved Users'
        for user in users:
            if user.selfie_approved == 'yes':
                userList.append(user)
    elif filtered == 'rejected_users':
        docName = 'Photo-Rejected Users'
        for user in users:
            if user.selfie_approved == 'no':
                userList.append(user)

    import openpyxl
    from openpyxl.utils import get_column_letter

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + docName + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Providers"

    row_num = 0

    columns = [
        (u"Name", 30),
        (u"Email", 40),
        (u"Age", 15),
        (u"Gender", 15),
        (u"Location", 40),
        (u"Membership Type", 20),
        (u"Expire At", 30),
        (u"Max Dates", 15),
    ]

    my_color = openpyxl.styles.colors.Color(rgb='FF99EDFC')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    from openpyxl.styles import Alignment

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    row_num = 1
    for user in userList:
        row_num = row_num + 1

        c = ws.cell(row=row_num, column=1)
        c.value = user.name

        c = ws.cell(row=row_num, column=2)
        c.value = user.email

        c = ws.cell(row=row_num, column=3)
        c.value = user.age

        c = ws.cell(row=row_num, column=4)
        c.value = user.gender

        c = ws.cell(row=row_num, column=5)
        c.value = user.address

        plans = Plan.objects.all()
        plan = plans[0]
        expired = ''
        dates = plan.dates
        membership_type = 'Bronze member'

        if user.premium:
            premiumStr = user.premium
            expired, dates = premiumStr.split("_")
            expired = time.strftime('%Y-%m-%d %H:%M %p', time.gmtime(int(expired) / 1000.0))
            membership_type = 'Premium member'

        c = ws.cell(row=row_num, column=6)
        c.value = membership_type

        c = ws.cell(row=row_num, column=7)
        c.value = expired

        c = ws.cell(row=row_num, column=8)
        c.value = dates

    wb.save(response)
    return response


def control(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    members = Member.objects.all().order_by('-id')
    members = getmembers(members)
    plan = Plan.objects.all()[0]
    return render(request, 'dating/control.html', {'users': members, 'plan':plan})

def deactivated(request, member_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    member = Member.objects.get(id=member_id)
    member.password = 'deactivated'
    member.save()
    members = Member.objects.all().order_by('-id')
    members = getmembers(members)
    plan = Plan.objects.all()[0]
    return render(request, 'dating/control.html', {'member_id': member_id, 'users': members, 'plan':plan})

def activated(request, member_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    member = Member.objects.get(id=member_id)
    member.password = ''
    member.save()
    members = Member.objects.all().order_by('-id')
    members = getmembers(members)
    plan = Plan.objects.all()[0]
    return render(request, 'dating/control.html', {'member_id': member_id, 'users': members, 'plan':plan})

def activated_users(request):
    request.session['filter'] = 'activated_users'
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    users = Member.objects.all().order_by('-id')
    userList = []
    for user in users:
        if user.password == '':
            userList.append(user)
    users = getmembers(userList)
    plan = Plan.objects.all()[0]
    return render(request, 'dating/control.html', {'users': users, 'filterF':'true', 'plan':plan})

def deactivated_users(request):
    request.session['filter'] = 'deactivated_users'
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return render(request, 'dating/login.html')
    users = Member.objects.all().order_by('-id')
    userList = []
    for user in users:
        if user.password == 'deactivated':
            userList.append(user)
    users = getmembers(userList)
    plan = Plan.objects.all()[0]
    return render(request, 'dating/control.html', {'users': users, 'filterF':'true', 'plan':plan})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def userupdate(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id', None)
        plan = request.POST.get('plan', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return render(request, 'dating/login.html')
        user = Member.objects.get(id=user_id)
        ps = Plan.objects.all()
        p = ps[0]
        if user is not None:
            if plan == '0':
                user.premium = ''
            elif plan == '1':
                expired = int(round(time.time() * 1000)) + int(p.months1) * 86400000 * 30
                dates = p.dates1
                user.premium = str(expired) + '_' + '1'
            elif plan == '2':
                expired = int(round(time.time() * 1000)) + int(p.months2) * 86400000 * 30
                dates = p.dates2
                user.premium = str(expired) + '_' + '2'
            elif plan == '3':
                expired = int(round(time.time() * 1000)) + int(p.months3) * 86400000 * 30
                dates = p.dates3
                user.premium = str(expired) + '_' + '3'
            user.phone = 'updated'
            user.save()
            members = Member.objects.all().order_by('-id')
            members = getmembers(members)
            plan = Plan.objects.all()[0]
            return render(request, 'dating/control.html', {'member_id': user_id, 'users': members, 'plan': plan, 'note':'update'})

    else: return redirect('/control')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def okayUpdatedMembership(request):
    if request.method == 'POST':
        user_id = request.POST.get('me_id', None)
        user = Member.objects.get(id=user_id)
        user.phone = ''
        user.save()
        return JsonResponse({'result_code':'0'})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def questionupdate(request):
    if request.method == 'POST':
        q = request.POST.get('q', '')
        q_id = request.POST.get('qid', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return render(request, 'dating/login.html')
        dynamical = Dynamical.objects.get(id=q_id)
        context = {}
        if dynamical is not None:
            dynamical.dynamical_question = q
            dynamical.save()
            qs = Dynamical.objects.all().order_by('-id')
            context = {'questions': qs, 'note': 'update'}
        else:
            qs = Dynamical.objects.all().order_by('-id')
            context = {'questions': qs}
        return render(request, 'dating/dynamical.html', context)

    else:
        return redirect('/questions')

def mybids(request, folder_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return redirect('/loginpage')
    bids = Bid.objects.filter(user_id=idx, folder_id=folder_id).order_by('-id')
    return render(request, 'dating/mybids.html', {'bids': bids, 'folder_id':folder_id})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def savebid(request,bid_id, folder_id):
    if request.method == 'POST':
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return redirect('/loginpage')
        caption = request.POST.get('caption', '')
        text = request.POST.get('text', '')
        bid = Bid.objects.get(id=bid_id)
        bid.user_id = idx
        bid.folder_id = folder_id
        bid.caption = caption
        bid.text = text
        bid.save()
        return redirect('/mybids/' + folder_id)

def addbid(request, folder_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return redirect('/loginpage')
    bid = Bid()
    bid.user_id = idx
    bid.folder_id = folder_id
    bid.save()
    return redirect('/mybids/' + folder_id)

def deletebid(request, bid_id, folder_id):
    Bid.objects.get(id=bid_id).delete()
    return redirect('/mybids/' + folder_id)

def loginpage(request):
    return render(request, 'dating/userlogin.html')

def signuppage(request):
    return render(request, 'dating/usersignup.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def userlogin(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        user = authenticate(username=email, password=password)
        if user is not None:
            login(request, user)
            return redirect('/folderhome')
        else:
            return render(request, 'dating/result.html',
                          {'result': 'You haven\'t been registered. Please sign up.'})
    else:
        return redirect('/loginpage')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def usersignup(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        user = User()
        user.username = email
        user.email = email
        user.password = password
        user.set_password(password)
        user.save()
        user = authenticate(username=email, password=password)
        if user is not None:
            login(request, user)
            return redirect('/folderhome')
        else:
            return render(request, 'dating/result.html',
                          {'result': 'You haven\'t been registered. Try again.'})
    else:
        return redirect('/signuppage')

def folderhome(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return redirect('/loginpage')
    folders = Folder.objects.filter(user_id=idx).order_by('-id')
    return render(request, 'dating/folderhome.html', {'folders':folders})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def createnewfolder(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return redirect('/loginpage')
        folder = Folder()
        folder.user_id = idx
        if len(name) == 0:
            name = 'New folder'
        folders = Folder.objects.filter(name=name)
        if folders.count() > 0:
            folders = Folder.objects.filter(name__startswith=name + '(')
            folder.name = name + '(' + str(folders.count() + 2) + ')'
        else:
            folders = Folder.objects.filter(name__startswith=name + '(')
            if folders.count() > 0:
                folder.name = name + '(' + str(folders.count() + 2) + ')'
            else:
                folder.name = name
        folder.files = '0'
        folder.save()
        return redirect('/folderhome')

def openfolder(request, folder_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return redirect('/loginpage')
    files = File.objects.filter(folder_id=folder_id).order_by('-id')
    folder_name = Folder.objects.get(id=folder_id).name
    return render(request, 'dating/filehome.html', {'folder_id':folder_id, 'folder_name':folder_name, 'files':files})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def addfile(request):
    if request.method == 'POST':
        folder_id = request.POST.get('folder_id', '')
        file = request.FILES['file']

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return redirect('/loginpage')

        f = File()
        files = File.objects.filter(name=file.name)
        if files.count() > 0:
            files = File.objects.filter(name__startswith=file.name + '(')
            f.name = file.name + '(' + str(files.count() + 2) + ')'
        else:
            files = File.objects.filter(name__startswith=file.name + '(')
            if files.count() > 0:
                f.name = file.name + '(' + str(files.count() + 2) + ')'
            else:
                f.name = file.name

        fs = FileSystemStorage()
        filename = fs.save(f.name, file)
        uploaded_file_url = fs.url(filename)

        f.folder_id = folder_id
        f.link = settings.URL + uploaded_file_url
        f.save()
        folder = Folder.objects.get(id=folder_id)
        folder.files = str(int(folder.files) + 1)
        folder.save()
        return redirect('/openfolder/' + folder_id + '/')

def deletefile(request, folder_id, file_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return redirect('/loginpage')
    file = File.objects.get(id=file_id)
    file.delete()
    fs = FileSystemStorage()
    fs.delete(file.link.replace(settings.URL + '/media/', ''))
    folder = Folder.objects.get(id=folder_id)
    folder.files = str(int(folder.files) - 1)
    folder.save()
    return redirect('/openfolder/' + folder_id + '/')

def filelogout(request):
    auth.logout(request)
    cache.clear()
    request.session.flush()
    request.user = AnonymousUser
    for key in request.session.keys():
        del request.session[key]
    return redirect('/loginpage')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def savefolder(request, folder_id):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        folder = Folder.objects.get(id=folder_id)
        if len(name) == 0:
            name = 'New folder'
        folders = Folder.objects.filter(name=name)
        if folders.count() > 0:
            folders = Folder.objects.filter(name__startswith=name + '(')
            folder.name = name.replace(find_between(name, '(', ')'), '') + '(' + str(folders.count() + 2) + ')'
        else:
            folders = Folder.objects.filter(name__startswith=name + '(')
            if folders.count() > 0:
                folder.name = name.replace(find_between(name, '(', ')'), '') + '(' + str(folders.count() + 2) + ')'
            else:
                folder.name = name
        folder.save()
        return redirect('/folderhome')
    else:
        return redirect('/folderhome')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def deletefolder(request, folder_id):
    if request.method == 'POST':
        folder = Folder.objects.get(id=folder_id)
        folder.delete()
        fs = FileSystemStorage()
        files = File.objects.filter(folder_id=folder_id)
        for file in files:
            file.delete()
            fs.delete(file.link.replace(settings.URL + '/media/', ''))
        bids = Bid.objects.filter(folder_id=folder_id)
        for bid in bids:
            bid.delete()
        return redirect('/folderhome')
    else:
        return redirect('/folderhome')

def find_between( s, first, last ):
    try:
        start = s.index( first ) + len( first ) - 1
        end = s.index( last, start ) + 1
        return s[start:end]
    except ValueError:
        return ""

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload_picture(request):

    if request.method == 'POST':

        image = request.FILES['file']
        member_id = request.POST.get('member_id')

        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)

        pictures = Picture.objects.filter(member_id=member_id)
        if pictures.count() >= 4:
            resp = {'result_code': '1'}
            return HttpResponse(json.dumps(resp))
        picture = Picture()
        picture.picture_url = settings.URL + uploaded_file_url
        picture.member_id = member_id
        picture.save()

        pictures = Picture.objects.filter(member_id=member_id).order_by('-id')
        serializer = PictureSerializer(pictures, many=True)
        resp = {'result_code': '0', 'pictures': serializer.data}
        return HttpResponse(json.dumps(resp))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updatePicture(request):

    if request.method == 'POST':

        image = request.FILES['file']
        picture_id = request.POST.get('picture_id')

        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)

        picture = Picture.objects.get(id=picture_id)
        picture.picture_url = settings.URL + uploaded_file_url
        picture.save()

        pictures = Picture.objects.filter(member_id=picture.member_id).order_by('-id')
        serializer = PictureSerializer(pictures, many=True)
        resp = {'result_code': '0', 'pictures': serializer.data}
        return HttpResponse(json.dumps(resp))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def deletePicture(request):

    if request.method == 'POST':
        picture_id = request.POST.get('picture_id')

        picture = Picture.objects.get(id=picture_id)
        picture.delete()
        fs = FileSystemStorage()
        fs.delete(picture.picture_url.replace(settings.URL + '/media/', ''))
        pictures = Picture.objects.filter(member_id=picture.member_id).order_by('-id')
        serializer = PictureSerializer(pictures, many=True)
        resp = {'result_code': '0', 'pictures': serializer.data}
        return HttpResponse(json.dumps(resp))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getPictures(request):

    if request.method == 'POST':
        member_id = request.POST.get('member_id')

        pictures = Picture.objects.filter(member_id=member_id).order_by('-id')
        serializer = PictureSerializer(pictures, many=True)
        resp = {'result_code': '0', 'pictures': serializer.data}
        return HttpResponse(json.dumps(resp))

    elif request.method == 'GET':
        pass

def mynotes(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return redirect('/loginpage')
    notes = Note.objects.filter(user_id=idx).order_by('-id')
    return render(request, 'dating/notebook.html', {'notes': notes})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def savenote(request,note_id):
    if request.method == 'POST':
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if idx is None:
            return redirect('/loginpage')
        caption = request.POST.get('caption', '')
        text = request.POST.get('text', '')
        note = Note.objects.get(id=note_id)
        note.user_id = idx
        note.caption = caption
        note.text = text
        note.save()
        return redirect('/mynotes')

def addnote(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    if idx is None:
        return redirect('/loginpage')
    note = Note()
    note.user_id = idx
    note.save()
    return redirect('/mynotes')

def deletenote(request, note_id):
    Note.objects.get(id=note_id).delete()
    return redirect('/mynotes')

def videochat(request):
    return render(request, 'dating/videochatview.html')

from django.core.mail import EmailMultiAlternatives

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendMail(request):
    if request.method == 'POST':
        toemail = 'buldier@gmail.com'
        fromemail = request.POST.get('email', None)
        name = request.POST.get('name', None)
        company = request.POST.get('company', None)
        body = request.POST.get('content', None)

        fromaddress = fromemail
        toaddress = toemail

        # html_content = ''
        body = 'Name: ' + name + '\n' + 'Company: ' + company + '\n\n' + body

        msg = EmailMultiAlternatives('Hi Admin', body, fromaddress, [toaddress])
        # msg.attach_alternative(html_content, "text/html")
        msg.send(fail_silently=False)

        resp = {'result_code':'0'}
        return HttpResponse(json.dumps(resp))

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def saveSettings(request):
    if request.method == 'POST':
        member_email = request.POST.get('member_email', '')
        max_age = request.POST.get('max_age', '')
        min_age = request.POST.get('min_age', '')
        gender = request.POST.get('gender', '')
        profile = request.POST.get('profile', '')

        ss = Settings.objects.filter(member_email=member_email)
        s = None
        if ss.count() == 0:
            s = Settings()
        else:
            s = ss[0]
        s.member_email = member_email
        s.max_age = max_age
        s.min_age = min_age
        s.gender = gender
        s.profile = profile
        s.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp))

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getSettings(request):
    if request.method == 'POST':
        member_email = request.POST.get('member_email', '')
        ss = Settings.objects.filter(member_email=member_email)
        if ss.count() > 0:
            s = ss[0]
            data = {
                'member_email':s.member_email,
                'max_age':s.max_age,
                'min_age':s.min_age,
                'gender':s.gender,
                'profile':s.profile
            }
            resp = {'result_code': '0', 'data':data}
            return HttpResponse(json.dumps(resp))
        else:
            resp = {'result_code': '1'}
            return HttpResponse(json.dumps(resp))









